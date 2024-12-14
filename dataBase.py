import openpyxl
import bisect
import os
import csv
from tkinter import Tk, messagebox, filedialog, Label, Entry, Button, ttk


class ExcelDatabase:
    def __init__(self, file_name):
        self.file_name = file_name
        self.key_index = {}
        self.sorted_keys = []
        if not os.path.exists(file_name):
            # Создаем файл, если он не существует
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["ID", "Name", "Author", "Cost"])  # Заголовки столбцов
            wb.save(file_name)
            wb.close()
        self.build_index()

    #постройка идексов для быстрого поиска по id
    def build_index(self):
        self.key_index.clear()  # Очистка предыдущего индекса
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            key = row[0].value
            if key is not None:
                self.key_index[key] = row[0].row  # Хранение только ключей и их строк
        wb.close()

    #добавляем запись в бд (сложность по сути O(1), если вставляем записи по возрастающему id)
    def add_record(self, record):
        key = record[0]
        if key in self.key_index:
            raise ValueError("Key already exists.")
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb.active
        ws.append(record)
        wb.save(self.file_name)
        wb.close()
        self.key_index[key] = ws.max_row
        bisect.insort(self.sorted_keys, key)

    #удаление записи только по id (не используется, оставил, если буду дорабатывать)  
    def delete_record_by_key(self, key):
        if key not in self.key_index:
            raise ValueError("Key not found.")
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb.active
        row_number = self.key_index[key]
        ws.delete_rows(row_number)
        wb.save(self.file_name)
        wb.close()
        self.build_index()
        
    # Удаление записей по значению заданного поля (не используется, оставил, если буду дорабатывать)  
    def delete_records_by_field(self, field_name, field_value):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb.active

        # Определяем номер столбца для заданного поля
        header = [cell.value for cell in ws[1]]
        if field_name not in header:
            raise ValueError(f"Поле '{field_name}' не найдено.")
        column_index = header.index(field_name) + 1

        rows_to_delete = []  # Список строк для удаления
        for row in ws.iter_rows(min_row=2):
            cell_value = row[column_index - 1].value
            if str(cell_value) == str(field_value):  # Сравниваем строки для универсальности
                rows_to_delete.append(row[0].row)

        if not rows_to_delete:
            raise ValueError(f"Записи с '{field_name} = {field_value}' не найдены.")

        # Удаляем строки в обратном порядке, чтобы индексы не сбивались
        for row_idx in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_idx)

        wb.save(self.file_name)
        wb.close()
        self.build_index()
        
    # Удаление записи по ключевому или неключевому полю
    def delete_record_auto(self, field_value):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb.active

        rows_to_delete = []

        if field_value in self.key_index:  # Если это ID
            row_number = self.key_index[field_value]
            rows_to_delete.append(row_number)
        else:  # Ищем по неключевому полю
            for row in ws.iter_rows(min_row=2):
                if any(str(cell.value) == str(field_value) for cell in row):
                    rows_to_delete.append(row[0].row)

        if not rows_to_delete:
            raise ValueError(f"Записи с '{field_value}' не найдены.")

        # Удаляем строки в обратном порядке
        for row_idx in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_idx)

        wb.save(self.file_name)
        wb.close()
        self.build_index()

    #очистка бд
    def delete_all_records(self):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb.active
        for row in range(2, ws.max_row + 1):
            ws.delete_rows(2)
        wb.save(self.file_name)
        wb.close()
        self.build_index()

    #поиск по id (вроде бы как o(1))
    def search_by_key(self, key):
        if key not in self.key_index:
            return None
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb.active
        row_number = self.key_index[key]
        record = [cell.value for cell in ws[row_number]]
        wb.close()
        return record

    #поиск по другим полям (линейная сложность)
    def search_by_non_key(self, field_value):
        matching_records = []
        try:
            field_value = float(field_value)
            wb = openpyxl.load_workbook(self.file_name)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Сравниваем только цену (индекс 3) с искомым значением
                if isinstance(row[3], (int, float)) and row[3] == field_value:
                    matching_records.append(row)
            wb.close()
        except ValueError:
            # Если значение не число, ищем по другим полям (Name, Author)
            wb = openpyxl.load_workbook(self.file_name)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(field_value == str(cell) for cell in row):
                    matching_records.append(row)
            wb.close()
        
        return matching_records

    #редактирование записи (используется id)
    def edit_record(self, key, new_record):
        if key not in self.key_index:
            raise ValueError("Key not found.")
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb.active
        row_number = self.key_index[key]
        for col, value in enumerate(new_record, start=1):
            ws.cell(row=row_number, column=col).value = value
        wb.save(self.file_name)
        wb.close()
        self.build_index()

    #резервное копирование
    def backup(self, backup_file_name):
        wb = openpyxl.load_workbook(self.file_name)
        wb.save(backup_file_name)
        wb.close()

    #восстановление из бэкапа
    def restore_from_backup(self, backup_file_name):
        if not os.path.exists(backup_file_name):
            raise FileNotFoundError("Backup file does not exist.")
        wb = openpyxl.load_workbook(backup_file_name)
        wb.save(self.file_name)
        wb.close()
        self.build_index()

    #экспорт бд в csv (есть проблемы с кодировкой)    
    def export_to_txt(self, txt_filename):
        # Открываем Excel-файл
        # Открываем Excel-файл
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb.active

        # Создаем текстовый файл с правильной кодировкой
        with open(txt_filename, mode='w', encoding='utf-8') as f:
            for row in ws.iter_rows(values_only=True):
                # Преобразуем каждую строку в строку текста, разделенную табуляцией
                row_text = '\t'.join(map(str, row))  # Используем табуляцию или любой другой разделитель
                f.write(row_text + '\n')  # Переход на новую строку

        wb.close()


class GUI:
    def __init__(self, root, db):
        self.root = root
        self.db = db
        self.root.title("база данных")
        self.root.geometry("800x400")

        #поля для ввода данных
        Label(root, text="ID").grid(row=0, column=0)
        Label(root, text="Name").grid(row=0, column=1)
        Label(root, text="Author").grid(row=0, column=2)
        Label(root, text="Cost").grid(row=0, column=3)

        self.entries = [Entry(root) for _ in range(4)]
        for i, entry in enumerate(self.entries):
            entry.grid(row=1, column=i)

        #кнопки, но тут названия говорящие
        Button(root, text="Добавить", command=self.add_record).grid(row=2, column=0)
        Button(root, text="Поиск", command=self.search_record).grid(row=2, column=1)
        Button(root, text="Удалить", command=self.delete_record).grid(row=2, column=2)
        Button(root, text="Редактировать", command=self.edit_record).grid(row=2, column=3)
        Button(root, text="Резервное копирование", command=self.backup_db).grid(row=3, column=0)
        Button(root, text="Восстановление", command=self.restore_db).grid(row=3, column=1)
        Button(root, text="Экспорт в TXT", command=self.export_db_to_txt).grid(row=3, column=2)
        Button(root, text="Удалить все записи", command=self.delete_all_records).grid(row=3, column=3)

        #табличка с данными под кнопками
        self.tree = ttk.Treeview(root, columns=("ID", "Name", "Author", "Cost"), show="headings")
        self.tree.heading("ID", text="ID")
        self.tree.heading("Name", text="Name")
        self.tree.heading("Author", text="Author")
        self.tree.heading("Cost", text="Cost")
        self.tree.grid(row=5, column=0, columnspan=4)

        self.load_table()

    #загрузка данных в бд для работы
    def load_table(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        wb = openpyxl.load_workbook(self.db.file_name)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            self.tree.insert("", "end", values=row)
        wb.close()

    #добавляем запись 
    def add_record(self):
        try:
            record = [entry.get() for entry in self.entries]
            record[0] = int(record[0])  #проверка id на инт 
            record[3] = float(record[3]) #проверка цены на флоат
            self.db.add_record(record)
            self.load_table()
            #messagebox.showinfo("Успех", "Запись добавлена!") #добавить при надобности
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    #поиск 
    def search_record(self):
        try:
            search_value = self.entries[0].get() or self.entries[1].get() or self.entries[2].get() or self.entries[3].get()  #что-нибудь из введенных данных
            matching_records = []

            if self.entries[0].get().isdigit():  # Поиск по id
                matching_records.append(self.db.search_by_key(int(search_value)))
            else:
                matching_records = self.db.search_by_non_key(search_value)

            #очистка предыдущих результатов поиска (очистка таблицы для вывода данных)
            for row in self.tree.get_children():
                self.tree.delete(row)

            if matching_records:
                #выводим найденные записи в Treeview
                for record in matching_records:
                    if record:  #чтобы исключить None
                        self.tree.insert("", "end", values=record)
                #messagebox.showinfo("Результат поиска", f"Найдено записей: {len(matching_records)}") #есть ошибка: некорректный вывод количества записей, которые нашли, при их отсутствии 
            else:
                messagebox.showinfo("Результат поиска", "Записи не найдены!")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    # Удаление записи по ID или неключевому полю
    def delete_record(self):
        try:
            id_value = self.entries[0].get()  # Значение ID
            other_field_value = self.entries[1].get() or self.entries[2].get() or self.entries[3].get()

            if id_value:  # Если введено значение ID
                self.db.delete_record_auto(int(id_value))
            elif other_field_value:  # Если ID пуст, ищем по другому полю
                self.db.delete_record_auto(other_field_value)
            else:
                raise ValueError("Не указано значение для удаления.")

            self.load_table()
            messagebox.showinfo("Успех", "Записи удалены!")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    #редактирование записи
    def edit_record(self):
        try:
            key = int(self.entries[0].get())
            new_record = [entry.get() for entry in self.entries]
            new_record[0] = int(new_record[0])
            new_record[3] = float(new_record[3])
            self.db.edit_record(key, new_record)
            self.load_table()
            #messagebox.showinfo("Успех", "Запись обновлена!") #добавить при надобности
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    #резервное копирование 
    def backup_db(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            self.db.backup(file_path)
            messagebox.showinfo("Успех", "Резервная копия создана!")

    #восстановление бд из бэкапа
    def restore_db(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            self.db.restore_from_backup(file_path)
            self.load_table()
            messagebox.showinfo("Успех", "База данных восстановлена!")

    #экспорт бд в csv
    def export_db_to_txt(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("TXT Files", "*.txt")])#(defaultextension=".csv", filetypes=[("CSV Files", "*.csv")])
        if file_path:
            self.db.export_to_txt(file_path)
            messagebox.showinfo("Успех", "Данные экспортированы в TXT!")

    #очистка бд
    def delete_all_records(self):
        try:
            self.db.delete_all_records()
            self.load_table()
            messagebox.showinfo("Успех", "Все записи удалены!")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))


if __name__ == "__main__":
    db = ExcelDatabase("database.xlsx")
    root = Tk()
    gui = GUI(root, db)
    root.mainloop()
